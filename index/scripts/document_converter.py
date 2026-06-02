import json
import os
import sys
import argparse
import logging
from pdf2image import convert_from_path
from pptx import Presentation
from openpyxl import load_workbook
from PIL import Image
import requests
from io import BytesIO
import base64
from dotenv import load_dotenv
import uuid
from docx2pdf import convert
import pdfkit
import shutil


# Global variable for describe_image batch size
DESCRIBE_IMAGE_BATCH_SIZE = 6

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def save_temp_image(image_data, temp_dir, file_extension='.png'):
    os.makedirs(temp_dir, exist_ok=True)
    temp_file_path = os.path.join(temp_dir, f"{uuid.uuid4()}{file_extension}")
    
    if isinstance(image_data, bytes):
        with open(temp_file_path, 'wb') as f:
            f.write(image_data)
    elif isinstance(image_data, Image.Image):
        image_data.save(temp_file_path)
    elif isinstance(image_data, BytesIO):
        with open(temp_file_path, 'wb') as f:
            f.write(image_data.getvalue())
    else:
        raise ValueError("Unsupported image_data type")
    
    return temp_file_path

def load_env(root_dir):
    env_path = os.path.join(root_dir, '.env')
    if not os.path.exists(env_path):
        logging.error(f".env file not found in {root_dir}")
        raise FileNotFoundError(f".env file not found in {root_dir}")
    
    load_dotenv(env_path)
    
    api_key = os.getenv('GRAPHRAG_API_KEY')
    if not api_key:
        logging.error("GRAPHRAG_API_KEY not found in .env file")
        raise ValueError("GRAPHRAG_API_KEY not found in .env file")
    
    logging.info("Environment variables loaded successfully")
    return api_key

# note, other than being super hacky, would be cheaper to upload the images and reference them in the request
# this is because base64 is just large by nature and model can't do anything to reduce tokens
def describe_images(image_data_list, api_key):
    global DESCRIBE_IMAGE_BATCH_SIZE
    try:
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}"
        }

        all_descriptions = []

        for i in range(0, len(image_data_list), DESCRIBE_IMAGE_BATCH_SIZE):
            batch = image_data_list[i:i+DESCRIBE_IMAGE_BATCH_SIZE]
            content = [
                {
                    "type": "text",
                    "text": ''' 
                        You are the leader of the enterprise architecture team, both technical and business architecture.
                        You are responsible for translating the content of the images into a clear, concise, and accurate message.
                        Provide a clear and concise title/explanation of the main ideas presented in each image. 
                        Use markdown formatting for the title and explanation, separating each image title/explanations with a horizontal rule (***).
                        Explain the key points, ideas, or information presented in the content of each image from the context of a subject matter expert in that topic/field. 
                        Use appropriate terminology and provide insightful analysis where relevant.
                        Avoid references to the image itself: Do not use phrases like "This image shows..." or "In this slide...". Instead, focus on explaining the content directly.
                        It is imperative that all critical details are maintained within the explanation - do not omit any details.
                    '''
                }
            ]

            for img_data in batch:
                if isinstance(img_data, str):  # If it's a file path
                    with Image.open(img_data) as img:
                        buffered = BytesIO()
                        img.save(buffered, format="PNG")
                        img_base64 = base64.b64encode(buffered.getvalue()).decode('utf-8')
                else:  # If it's already a BytesIO object
                    img_base64 = base64.b64encode(img_data.getvalue()).decode('utf-8')
                
                content.append({
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:image/png;base64,{img_base64}"
                    }
                })

            logging.info(f"Sending batch of {len(batch)} images to OpenAI for description")
            payload = {
                "model": "gpt-4o-mini",
                "messages": [
                    {
                        "role": "user",
                        "content": content
                    }
                ],
                "max_tokens": 1000 * len(batch)
            }

            response = requests.post("https://api.openai.com/v1/chat/completions", headers=headers, json=payload)
            result = response.json()
            logging.info("Image descriptions received from OpenAI")

            if 'choices' in result and len(result['choices']) > 0:
                descriptions = result['choices'][0]['message']['content'].split('***')
                all_descriptions.extend(descriptions)
            else:
                logging.error(f"Unexpected API response: {json.dumps(result, indent=2)}")
                all_descriptions.append("Error: Unexpected API response format")

        return all_descriptions
    except Exception as e:
        logging.error(f"Error describing images: {str(e)}")
        return f"Error describing images: {str(e)}"
    
def convert_image(file_path, api_key):
    return describe_images([file_path], api_key)
    
def convert_pdf(file_path, temp_dir, api_key):

    images = convert_from_path(file_path)
    image_paths = []
    
    for i, image in enumerate(images, 1):
        # Save the page as an image
        img_path = save_temp_image(image, temp_dir, f'page_{i}.png')
        image_paths.append(img_path)
    
    # Describe all page images
    descriptions = describe_images(image_paths, api_key)
    descriptions = '\n\n'.join(f"<BEGIN PAGE {i+1}>\n{page.strip()}\n</END PAGE {i+1}>" for i, page in enumerate(descriptions))

    # Clean up temporary image files
    for img_path in image_paths:
        os.remove(img_path)
    
    return descriptions

def convert_ppt(file_path, temp_dir, api_key):
    # Convert PPT to PDF
    pdf_path = os.path.join(temp_dir, f"{uuid.uuid4()}.pdf")
    prs = Presentation(file_path)
    prs.save(pdf_path)
    
    # Use the existing convert_pdf function
    result = convert_pdf(pdf_path, temp_dir, api_key)
    
    # Remove the temporary PDF file
    os.remove(pdf_path)
    
    return result

def convert_doc(file_path, temp_dir, api_key):    

    # Convert DOCX to PDF
    pdf_path = os.path.join(temp_dir, f"{uuid.uuid4()}.pdf")
    convert(file_path, pdf_path)
    
    # Use the existing convert_pdf function
    result = convert_pdf(pdf_path, temp_dir, api_key)
    
    # Remove the temporary PDF file
    os.remove(pdf_path)
    
    return result

def convert_html(file_path, temp_dir, api_key):
 
    # Convert HTML to PDF
    pdf_path = os.path.join(temp_dir, f"{uuid.uuid4()}.pdf")
    
    # Use pdfkit to convert HTML to PDF
    pdfkit.from_file(file_path, pdf_path)
    
    # Use the existing convert_pdf function
    result = convert_pdf(pdf_path, temp_dir, api_key)
    
    # Remove the temporary PDF file
    os.remove(pdf_path)
    
    return result

def convert_xls(file_path, api_key):
    wb = load_workbook(file_path)
    text_content = []
    for sheet_index, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]
        text_content.append(f"Sheet {sheet_index}: {sheet_name}")
        
        for row_index, row in enumerate(ws.iter_rows(values_only=True), 1):
            row_content = ' | '.join(str(cell) for cell in row if cell is not None)
            if row_content.strip():
                text_content.append(f"Row {row_index}: {row_content}")
        
        text_content.append("")  # Add a blank line between sheets
    
    # Check for charts and images
    for sheet in wb:
        image_paths = []
        for image in sheet._images:
            img_path = save_temp_image(image._data())
            image_paths.append(img_path)
        
        if image_paths:
            descriptions = describe_images(image_paths, api_key)
            for i, description in enumerate(descriptions.split('\n\n')):
                text_content.append(f"[Image {i+1} in {sheet.title}: {description.strip()}]")
            
            # Clean up temporary image files
            for img_path in image_paths:
                os.remove(img_path)
    
    return '\n'.join(text_content)


def convert_to_text(file_path, temp_dir, api_key):
    _, file_extension = os.path.splitext(file_path.lower())

    conversion_functions = {
        '.ppt': convert_ppt,
        '.pptx': convert_ppt,
        '.pdf': lambda path, temp, key: convert_pdf(path, temp, key),
        '.doc': convert_doc,
        '.docx': convert_doc,
        '.xls': convert_xls,
        '.xlsx': convert_xls,
        '.html': convert_html,
        '.jpg': convert_image,
        '.jpeg': convert_image,
        '.png': convert_image,
        '.gif': convert_image
    }

    if file_extension in ['.txt', '.md', '.csv']:
        with open(file_path, 'r', encoding='utf-8') as f:
            return f.read()

    convert_func = conversion_functions.get(file_extension)
    if convert_func:
        return convert_func(file_path, temp_dir, api_key)
    else:
        return "Unsupported file format"

def process_directory(input_dir, output_dir, temp_dir, api_key):
    for root, _, files in os.walk(input_dir):
        for file in files:
            input_file_path = os.path.join(root, file)
            relative_path = os.path.relpath(input_file_path, input_dir)
            output_file_path = os.path.join(output_dir, os.path.splitext(relative_path)[0] + '.txt')
            
            try:
                logging.info(f"Processing file: {input_file_path}")
                # Create necessary directories
                os.makedirs(os.path.dirname(output_file_path), exist_ok=True)
                
                _, file_extension = os.path.splitext(input_file_path.lower())
                
                if file_extension in ['.txt', '.md', '.csv']:
                    shutil.copy2(input_file_path, output_file_path)
                    logging.info(f"Copied {input_file_path} to {output_file_path}")
                else:
                    text_content = convert_to_text(input_file_path, temp_dir, api_key)
                    
                    if text_content == "Unsupported file format":
                        logging.warning(f"Skipping unsupported file: {input_file_path}")
                        continue
                    
                    with open(output_file_path, 'w', encoding='utf-8') as f:
                        f.write(text_content)
                    logging.info(f"Converted {input_file_path} to {output_file_path}")
            except Exception as e:
                logging.error(f"Error processing {input_file_path}: {str(e)}")

def main():
    try:
        parser = argparse.ArgumentParser(description="Convert various document formats to text files.")
        parser.add_argument("root_dir", nargs='?', default="index", help="Root directory containing .env file, input, output, and temp folders (default: index)")
        parser.add_argument("staging_dir", nargs='?', default="staging", help="Directory of the files to be converted (default: staging)")    
        parser.add_argument("output_dir", nargs='?', default="input", help="Directory to place converted files (default: input)")    
        parser.add_argument("temp_dir", nargs='?', default="temp", help="Directory for temporary files (default: temp)")    
        parser.add_argument("--batch-size", type=int, default=10, help="Batch size for describing images (default: 10)")
        args = parser.parse_args()

        global DESCRIBE_IMAGE_BATCH_SIZE
        DESCRIBE_IMAGE_BATCH_SIZE = args.batch_size

        root_dir = os.path.abspath(args.root_dir)
        input_dir = os.path.join(root_dir, args.staging_dir)
        output_dir = os.path.join(root_dir, args.output_dir)
        temp_dir = os.path.join(root_dir, args.temp_dir)

        logging.info(f"Root directory: {root_dir}")
        logging.info(f"Input directory: {input_dir}")
        logging.info(f"Output directory: {output_dir}")
        logging.info(f"Temp directory: {temp_dir}")

        # Validate directories
        for dir_path, dir_name in [(root_dir, "Root"), (input_dir, "Input")]:
            if not os.path.isdir(dir_path):
                logging.error(f"{dir_name} directory '{dir_path}' does not exist.")
                raise FileNotFoundError(f"{dir_name} directory '{dir_path}' does not exist.")

        os.makedirs(output_dir, exist_ok=True)
        os.makedirs(temp_dir, exist_ok=True)

        api_key = load_env(root_dir)
        process_directory(input_dir, output_dir, temp_dir, api_key)

        logging.info("Document conversion completed successfully")

    except FileNotFoundError as e:
        logging.error(f"File not found: {str(e)}")
        sys.exit(1)
    except ValueError as e:
        logging.error(f"Value error: {str(e)}")
        sys.exit(1)
    except Exception as e:
        logging.error(f"An unexpected error occurred: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    main()
